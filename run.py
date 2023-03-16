from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QDialog, QTableWidgetItem, QLabel, QPushButton
from PyQt5.QtCore import QSize, Qt, QThread, pyqtSignal
from PyQt5 import QtCore
from PyQt5.QtGui import QIcon, QPixmap, QCloseEvent
import csv
import logging
import os
from handle_form1 import Handle_Form1
from handle_main import Handle_Main
import openpyxl

from src.MDB.Product import Product
from src.MDB.APICaller import APICaller
from src.MDB.config import ConfigSMartTool
from src.MDB.Utils import LoadProductListSingleSheet
from src.MDB.api_download import download_image, download_product, download_text

# Global values
percent = 0
logger = logging.getLogger('main')
logger.setLevel(logging.INFO)
class UI():
  def __init__(self):
    self.form1 = None
    self.mainUI = QMainWindow()
    self.mainHandle = Handle_Main(self.mainUI)
    self.mainHandle.btn1_1.clicked.connect(lambda: self.selectFile())
    self.mainHandle.table_excel.cellChanged.connect(self.onCellChanged)
    self.mainHandle.btn1_4.clicked.connect(lambda: self.export_data_screen_1())
    self.mainHandle.btn1_4.clicked.connect(lambda: self.applyButtonStyle_main(self.mainHandle.btn2))
    
    self.mainHandle.btn1_2.clicked.connect(lambda: self.loadMainForm(0))
    self.mainHandle.btn1_3.clicked.connect(lambda: self.show123())
    # self.mainHandle.btn2_1.clicked.connect(self.read_csv)
    
    self.row = 1
    self.column = 1
    
    self.header = ["Jan code","Tên sản phẩm","Image","Day Update"]
    
    self.mainUI.show()
  
    for w in self.mainHandle.frame.findChildren(QPushButton):
        # Add click event listener
        w.clicked.connect(lambda _, button=w: self.applyButtonStyle_main(button)) 

    for w in self.mainHandle.page_1.findChildren(QPushButton):
        # Add click event listener
        w.clicked.connect(lambda _, button=w: self.applyButtonStyle_page_1(button)) 

    for w in self.mainHandle.page_2.findChildren(QPushButton):
        # Add click event listener
        w.clicked.connect(lambda _, button=w: self.applyButtonStyle_page_2(button)) 
    


  def applyButtonStyle_main(self, button):
      # Reset style for other buttons
      for w in self.mainHandle.frame.findChildren(QPushButton):
          # If the button name is not equal to clicked button name
           if w.objectName() != button.objectName(): 
              # Create default style by removing the left border
              # Lets remove the bottom border style
              print(w)
              if(len(button.objectName())<=4):
                defaultStyle = w.styleSheet().replace("background-color: rgb(202, 168, 107);", "background-color: rgb(110, 145, 223);")
              else:
                print(len(button.objectName()))
                defaultStyle = w.styleSheet().replace("background-color: rgb(201, 131, 106);", "background-color: rgb(202, 168, 107);")
                
              # # Lets also remove the left border style
              # defaultStyle = defaultStyle.replace("border-left: 2px solid  rgb(0, 136, 255);", "")

              # Apply the default style
              w.setStyleSheet(defaultStyle)

      # Apply new style to clicked button
      # Sender = clicked button
      # Get the clicked button stylesheet then add new left-border style to it
      # Lets add the bottom border style
      if(len(button.objectName())<=4):
        newStyle = button.styleSheet() + ("background-color: rgb(202, 168, 107);")
      else:
        newStyle = button.styleSheet() + ("background-color: rgb(201, 131, 106);")
      # Apply the new style
      button.setStyleSheet(newStyle)    
  def applyButtonStyle_page_1(self, button):
      # Reset style for other buttons
      for w in self.mainHandle.page_1.findChildren(QPushButton):
          # If the button name is not equal to clicked button name
           if w.objectName() != button.objectName(): 
              # Create default style by removing the left border
              # Lets remove the bottom border style
              defaultStyle = w.styleSheet().replace("background-color: rgb(201, 131, 106);", "background-color: rgb(202, 168, 107);")
                
              # # Lets also remove the left border style
              # defaultStyle = defaultStyle.replace("border-left: 2px solid  rgb(0, 136, 255);", "")

              # Apply the default style
              w.setStyleSheet(defaultStyle)

      # Apply new style to clicked button
      # Sender = clicked button
      # Get the clicked button stylesheet then add new left-border style to it
      # Lets add the bottom border style
      newStyle = button.styleSheet() + ("background-color: rgb(201, 131, 106);")
      # Apply the new style
      button.setStyleSheet(newStyle)    

  def applyButtonStyle_page_2(self, button):
      # Reset style for other buttons
      for w in self.mainHandle.page_2.findChildren(QPushButton):
          # If the button name is not equal to clicked button name
           if w.objectName() != button.objectName(): 
              # Create default style by removing the left border
              # Lets remove the bottom border style
              defaultStyle = w.styleSheet().replace("background-color: rgb(201, 131, 106);", "background-color: rgb(202, 168, 107);")
                
              # # Lets also remove the left border style
              # defaultStyle = defaultStyle.replace("border-left: 2px solid  rgb(0, 136, 255);", "")

              # Apply the default style
              w.setStyleSheet(defaultStyle)

      # Apply new style to clicked button
      # Sender = clicked button
      # Get the clicked button stylesheet then add new left-border style to it
      # Lets add the bottom border style
      newStyle = button.styleSheet() + ("background-color: rgb(201, 131, 106);")
      # Apply the new style
      button.setStyleSheet(newStyle) 


  def show123(self):
    self.form1 = QDialog()
    self.form1Handle = Handle_Form1(self.form1)
    self.form1Handle.setupUi(self.form1)
    self.form1.exec()
    
  def loadMainForm(self,data):

    # self.mainUI.hide()
    self.form1 = QDialog()
    # self.form1.setWindowFlag(Qt.FramelessWindowHint)
    self.form1Handle = Handle_Form1(self.form1)
    self.form1Handle.setupUi(self.form1)
    
    
    self.form1Handle.progressBar.setValue(0)
    self.form1Handle.label_2.setText("")
    self.calculationThread = Call_download()
    self.calculationThread.updateProgressBar.connect(self.updateProgressBarValue)
    self.calculationThread.label_jan_cd.connect(self.setTextLable)
    self.calculationThread.start()
    # self.form1.show()
    self.form1.exec()
    return
  
  def updateProgressBarValue(self, value):
    self.form1Handle.progressBar.setValue(value-1)
    if(value+1>100):
      self.form1.hide()
    print(value)
  def setTextLable(self, value):
    self.form1Handle.label_2.setText("Download data "+value+"...")
    print(value)
        
  def selectFile(self):
    dir = "D:\\Scropper\\python_app\\data"
    filters = "Excel files (*.xlsx)"
    selected_filter = "Excel (*.xlsx)"
    # options = "" # ???
    fname = QFileDialog.getOpenFileName(caption=" Open File ", directory=dir, filter=filters, initialFilter=selected_filter)    
    if len(fname[0]) ==0:
      return
    self.show_data_excel(fname[0])
    

   #xu ly excel
  # def get_cell_value_list(sheet):
  #   return [[cell.value for cell in row] for row in sheet] 
    
  def onCellChanged(self, row, column):
    item = self.mainHandle.table_excel.item(row, column)    
    boxcheck = item.checkState()
    if boxcheck == QtCore.Qt.CheckState.Checked:
      try:
        print("checked row: ",row)
        list = []
        for _col in range(1, self.column-1):
          item = self.mainHandle.table_excel.item(row, _col)
          list.append(item.text())
        print("row {}".format(row), str(list))
      except:
        print("no data")
    
    
  def show_data_excel(self, link):
    # link = self.form1Handle.lbl_link.toPlainText()
    self.wb = openpyxl.load_workbook(link)
    self.sheet = self.wb['Sheet1']
    
        # Đếm số hàng có giá trị
    count = 0
    for row in self.sheet.iter_rows():
        for cell in row:
            if cell.value:
                count += 1
    # sheet_1 =self.get_cell_value_list(self.sheet[0])
    # self.sheet = self.wb.active
    
    self.mainHandle.table_excel.setRowCount(count)
    self.mainHandle.table_excel.setColumnCount(int(self.sheet.max_column)+2)
    
    
    list_values = list(self.sheet.values)
    
    list_title = self.header
    list_title.append("Status")
    list_title.append("Select")
    
    self.mainHandle.table_excel.setHorizontalHeaderLabels(list_title)
    self.row = self.mainHandle.table_excel.rowCount()
    self.column = self.mainHandle.table_excel.columnCount()
    
    list_jan_cd_code = []
    
    
    row_index = 0
    for value_tuple in list_values[1:]:
      col_index = 0
      if value_tuple[0] is None:
        continue
      
      list_jan_cd_code.append(str(value_tuple[0]))
      # data = self.getDataMDB(str(value_tuple[0]),str(value_tuple[1]))
      # print(data)
      
      
      if "5" in str(value_tuple[1]):
        text = "update"
      else:
        text = "insert"
      data_jan_cd = self.read_csv(value_tuple[0])
      list_data_show_table = [data_jan_cd['shn_cd'],data_jan_cd['mkr_mei'],'',data_jan_cd['upd_date']]
      
      for value in list_data_show_table:
          self.mainHandle.table_excel.setColumnWidth(col_index, 100)
          self.mainHandle.table_excel.setRowHeight(row_index, 100)
          self.mainHandle.table_excel.setIconSize(QSize(100,100))
          # pixmap = QPixmap("D:\\Scropper\python_app\\result_files\\4527760825491\\4527760825491_0000000000000_0.JPG")
          # pixmap_resized = pixmap.scaled(100, 100, QtCore.Qt.KeepAspectRatio)
          item = QTableWidgetItem()   
          icon = QIcon(f'D:\\Scropper\python_app\\result_files\\{value_tuple[0]}\\{value_tuple[0]}_0000000000000_0.JPG')
          item.setIcon(icon)
          
          
          if col_index==2:
            self.mainHandle.table_excel.setItem(row_index, 2, item)
          else:
            self.mainHandle.table_excel.setItem(row_index , col_index, QTableWidgetItem(str(value)))
          col_index += 1
      #status
      status = QTableWidgetItem(text)
      self.mainHandle.table_excel.setItem(row_index,col_index, status)

      #check_box
      checkbox_item = QTableWidgetItem()
      checkbox_item.setCheckState(QtCore.Qt.Unchecked)
      self.mainHandle.table_excel.setItem(row_index,col_index+1, checkbox_item)
      checkbox_item.checkState
      
      row_index += 1
    # data = self.getDataMDB(list_jan_cd_code,'20000101')    
    # print(data)
    return
    
  def export_data_screen_1(self):
      
    self.data_screen_1 = []
    self.data_screen_1.append(self.header[:-1])
    for index_row in range(self.row):
      item = self.mainHandle.table_excel.item(index_row, self.column-1)    
      if item is None: continue
      boxcheck = item.checkState()
      if boxcheck == 2 :
        list = []
        for _col in range(self.column-1):
          data = self.mainHandle.table_excel.item(index_row, _col)
          list.append(data.text())
        self.data_screen_1.append(list)
    self.show_data_excel_2()
    print(self.data_screen_1)
    # print(self.row, self.column)
    
    
#screen2
  def show_data_excel_2(self):

    
    list_values = self.data_screen_1
    row_screen2 = len(list_values)-1
    column_screen2 = len(list_values[0])
    
    self.mainHandle.table_excel_2.setRowCount(row_screen2)
    self.mainHandle.table_excel_2.setColumnCount(column_screen2)
    self.mainHandle.table_excel_2.setHorizontalHeaderLabels(list_values[0])
    
    row_index = 0
    for value_tuple in list_values[1:]:
      col_index = 0
      if value_tuple[0] is None:
        continue 
      for value in value_tuple:
        
        self.mainHandle.table_excel_2.setColumnWidth(col_index, 100)
        self.mainHandle.table_excel_2.setRowHeight(row_index, 100)
        self.mainHandle.table_excel_2.setIconSize(QSize(100,100))
        
        item = QTableWidgetItem()   
        icon = QIcon(f'D:\\Scropper\python_app\\result_files\\{value_tuple[0]}\\{value_tuple[0]}_0000000000000_0.JPG')
        item.setIcon(icon)
      
        if col_index==2:
          self.mainHandle.table_excel_2.setItem(row_index, 2, item)
        else:
          self.mainHandle.table_excel_2.setItem(row_index , col_index, QTableWidgetItem(str(value)))
        col_index += 1        
      row_index += 1



  
  def read_csv(self, path_jan_cd):
    with open(f'D:\\Scropper\\python_app\\result_files\\{path_jan_cd}\\{path_jan_cd}.csv', encoding='utf-8') as csvfile:
      reader = csv.DictReader(csvfile)
      for row in reader:
           result = row
      
    return result


class Call_download(QThread):
  updateProgressBar = pyqtSignal(int)
  label_jan_cd = pyqtSignal(str)
  def run(self):
    self.path = 'D:\\Scropper\\python_app\\result_files'
    self.products = 'D:\\Scropper\\python_app\\data\\jan_cd.xlsx'
    self.sheet = 1 
    logger.info("Configuration")
    logger.info("Default from_date :{0}".format(
        ConfigSMartTool.APIConfig['from_date']))
    absPath = os.path.abspath(self.path)
    logger.info("Abs path :{0}".format(absPath))

    # Download
    logger.debug("Download")
    productListExcelFile = self.products
    logger.info("Product list file :{0}".format(productListExcelFile))

    # Load product list
    productList = LoadProductListSingleSheet(
        productListExcelFile, self.sheet-1)
    logger.info("Number of product :{0}".format(len(productList)))
    self.total_jan_cd = len(productList)
    # Set storage path
    for product in productList:
        product: Product = product
        product.storagePath = absPath

    # For each product
    apiCaller = APICaller()
    apiCaller.authen()
    download_result_list = []
    total_jan_cd = len(productList)
    percent = 1
    for product in productList:
      self.updateProgressBar.emit(int((percent)/(total_jan_cd)*100))
      self.label_jan_cd.emit(product.shn_cd)
      product: Product = product
      # product.describe()
      # If existing and not overwrite, skip
      logger.debug("Product path :{0}".format(product.getBasePath()))
      if product.isExistingData():
          logger.info(
              "{0} Existing, still download : ".format(product.shn_cd))
          download_result = download_product(
              product=product, apiCaller=apiCaller)
          download_result_list.append([product.shn_cd, download_result])
      else:
          logger.info(" {0} New downloading".format(product.shn_cd))
          download_result = download_product(
              product=product, apiCaller=apiCaller,
              )
          download_result_list.append([product.shn_cd, download_result])
      percent = percent+1
    return


if __name__ == "__main__":
  app = QApplication([])
  
  ui = UI()
  
  app.exec_()